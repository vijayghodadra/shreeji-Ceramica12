from __future__ import annotations

import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook
from fastapi import FastAPI, Query, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse

from extractor import (
    DEFAULT_CACHE_PATH,
    DEFAULT_IMAGES_DIR,
    DEFAULT_KOHLER_CACHE_PATH,
    AQUANT_VARIANT_COLOR_MAP,
    image_relative_path,
    normalize_code,
    normalize_text,
)
from runtime_paths import get_backend_base_dir

BASE_DIR = get_backend_base_dir()
KOHLER_CACHE_PATH = os.path.join(str(BASE_DIR), "kohler_cache.json")
KOHLER_EXCEL_PATH = os.path.join(str(BASE_DIR), "kohler_catalog_full.xlsx")
USE_CACHE = False
IMAGES_DIR = Path(DEFAULT_IMAGES_DIR)
IMAGES_DIR.mkdir(parents=True, exist_ok=True)
WINDOWS_FORBIDDEN_SEGMENT_CHARS = set('<>:"\\|?*')

KOHLER_NO_IMAGE_CODES = {
    normalize_code(code)
    for code in {
        "K-24740IN-K4",
        "K-24740IN-7",
        "K-705087IN-SHP",
        "K-704613IN-CP",
        "K-1286731",
        "K-1063956",
        "K-1060831",
        "K-1042534",
        "K-82958",
        "K-17663IN-0",
    }
}

CATALOG_SOURCES = {
    "aquant": {
        "label": "Aquant",
        "excel_path": BASE_DIR / "aquant_catalog_full.xlsx",
        "cache_path": Path(DEFAULT_CACHE_PATH),
    },
    "kohler": {
        "label": "Kohler",
        "excel_path": Path(KOHLER_EXCEL_PATH),
        "cache_path": Path(KOHLER_CACHE_PATH),
    },
}

FALLBACK_PRODUCTS_PATHS = [
    BASE_DIR / "products_complete.json",
    BASE_DIR / "products.json",
    BASE_DIR.parent / "backend" / "products_complete.json",
    BASE_DIR.parent / "backend" / "products.json",
    Path.cwd() / "products_complete.json",
    Path.cwd() / "products.json",
]

# Manual corrections requested from catalog PDF references.
PRODUCT_OVERRIDES = {
    "1961": {
        "name": "Semi Counter Basin",
        "price": 16500,
        "color": "White",
        "size": "700 x 470 x 210 mm",
        "image": "/images/1961SEMI.png",
    },
    "1962": {
        "name": "Three Hole Semi Counter Basin",
        "price": 17500,
        "color": "White",
        "size": "700 x 470 x 210 mm",
        "image": "/images/1962.png",
    },
    "1963ab": {
        "price": 59500,
    },
    "1963g": {
        "price": 59500,
    },
    "2031": {
        "name": "Concealed Wall Mounted Single Lever Dusch Mixer 15 mm inlet",
        "price": 8500,
        "details": "Concealed Wall Mounted Single Lever Dusch Mixer 15 mm inlet",
        "image": "/images/2031.png",
    },
    "4004": {
        "price": 5750,
    },
    "1505cp": {
        "price": 3590,
    },
    "1342cp": {
        "price": 23500,
    },
}

MANUAL_QUERY_RESULTS = {
    "1001ft": [
        {
            "code": "1001-FT",
            "variant": "FT",
            "color": "Chrome",
            "price": 1090,
            "image": "/images/1001.png",
            "name": "Pop-Up Waste Coupling",
            "details": "Pop-Up Waste Coupling",
            "size": "32 mm x 75 mm",
            "base_code": "1001",
        },
    ],
    "1001ht": [
        {
            "code": "1001-HT",
            "variant": "HT",
            "color": "Chrome",
            "price": 1090,
            "image": "/images/1001.png",
            "name": "Pop-Up Waste Coupling",
            "details": "Pop-Up Waste Coupling",
            "size": "32 mm x 75 mm",
            "base_code": "1001",
        },
    ],
    "1017cp": [
        {
            "code": "1017-CP",
            "variant": "CP",
            "color": "Chrome",
            "price": 1350,
            "image": "/images/1017CP.png",
            "name": "Pop-Up Waste Coupling",
            "details": "Pop-Up Waste Coupling",
            "size": "32 mm x 125 mm",
            "base_code": "1017",
        },
    ],
    "1018ft": [
        {
            "code": "1018-FT",
            "variant": "FT",
            "color": "Chrome",
            "price": 2150,
            "image": "/images/1001.png",
            "name": "Pop-Up Waste Coupling",
            "details": "Pop-Up Waste Coupling",
            "size": "32 mm x 175 mm",
            "base_code": "1018",
        },
    ],
    "1023": [
        {
            "code": "1023",
            "variant": None,
            "color": None,
            "price": 4950,
            "name": "Stone Pop-Up Waste Coupling",
            "details": "Stone Pop-Up Waste Coupling",
            "size": "32 mm x 120 mm",
            "base_code": "1023",
        },
    ],
    "1871": [
        {"code": "1871 AB + W", "variant": "AB+W", "color": "Antique Bronze", "price": 44500, "image": "/images/1857W.png", "name": "WC With PVC Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover White", "size": "760 x 440 x 750 mm", "base_code": "1871"},
        {"code": "1871 G + W", "variant": "G+W", "color": "Gold", "price": 44500, "image": "/images/1857W.png", "name": "WC With PVC Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover White", "size": "760 x 440 x 750 mm", "base_code": "1871"},
        {"code": "1871 AB + WN", "variant": "AB+WN", "color": "Antique Bronze", "price": 59500, "image": "/images/1857WN.png", "name": "WC With Wooden Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover Walnut Colour (Engineered Wood)", "size": "760 x 440 x 750 mm", "base_code": "1871"},
        {"code": "1871 G + WN", "variant": "G+WN", "color": "Gold", "price": 59500, "image": "/images/1857WN.png", "name": "WC With Wooden Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover Walnut Colour (Engineered Wood)", "size": "760 x 440 x 750 mm", "base_code": "1871"},
    ],
    "1871abw": [
        {"code": "1871 AB + W", "variant": "AB+W", "color": "Antique Bronze", "price": 44500, "image": "/images/1857W.png", "name": "WC With PVC Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover White", "size": "760 x 440 x 750 mm", "base_code": "1871"},
    ],
    "1871gw": [
        {"code": "1871 G + W", "variant": "G+W", "color": "Gold", "price": 44500, "image": "/images/1857W.png", "name": "WC With PVC Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover White", "size": "760 x 440 x 750 mm", "base_code": "1871"},
    ],
    "1871abwn": [
        {"code": "1871 AB + WN", "variant": "AB+WN", "color": "Antique Bronze", "price": 59500, "image": "/images/1857WN.png", "name": "WC With Wooden Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover Walnut Colour (Engineered Wood)", "size": "760 x 440 x 750 mm", "base_code": "1871"},
    ],
    "1871gwn": [
        {"code": "1871 G + WN", "variant": "G+WN", "color": "Gold", "price": 59500, "image": "/images/1857WN.png", "name": "WC With Wooden Seat Cover", "details": "1 Piece Floor Mounted Couple Suite S Trap Siphonic Flushing System, Trap Distance 300 mm Seat Cover Walnut Colour (Engineered Wood)", "size": "760 x 440 x 750 mm", "base_code": "1871"},
    ],
    "2652bg": [
        {"code": "2652 BG", "variant": "BG", "color": "Brushed Gold", "price": 49500, "image": "/images/2652BG.png", "name": "Brass & Glass Tall Basin Mixer With Visible Waterfall Flow", "details": "Brass & Glass Tall Basin Mixer With Visible Waterfall Flow", "size": None, "base_code": "2652"},
    ],
}

app = FastAPI(title="Multi Catalog Product Search API")


def _cors_origins_from_env() -> tuple[list[str], str | None, bool]:
    default_origins = [
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3001",
        "null",
    ]
    env_value = os.environ.get("CORS_ORIGINS", "").strip()
    origins = [item.strip() for item in env_value.split(",") if item.strip()] if env_value else default_origins

    if "*" in origins:
        # Star origin cannot be used with credentials in CORS middleware.
        return ["*"], None, False

    origin_regex = os.environ.get(
        "CORS_ORIGIN_REGEX",
        r"^https?://((localhost|127\.0\.0\.1)(:\d+)?|[a-z0-9-]+\.vercel\.app|[a-z0-9-]+\.onrender\.com)$",
    ).strip()
    return origins, origin_regex or None, True


_allow_origins, _allow_origin_regex, _allow_credentials = _cors_origins_from_env()

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allow_origins,
    allow_origin_regex=_allow_origin_regex,
    allow_credentials=_allow_credentials,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _candidate_images_dirs() -> list[Path]:
    candidates = [
        IMAGES_DIR,
        BASE_DIR / "dist" / "ProductCatalogBackend" / "_internal" / "images",
    ]

    resolved = []
    seen = set()
    for candidate in candidates:
        try:
            normalized = candidate.resolve()
        except OSError:
            continue
        if normalized in seen or not normalized.exists() or not normalized.is_dir():
            continue
        seen.add(normalized)
        resolved.append(normalized)

    return resolved or [IMAGES_DIR]


def _resolve_existing_image_path(relative_path: str) -> Path | None:
    clean_relative = image_relative_path(relative_path)
    if not clean_relative:
        return None

    for images_dir in _candidate_images_dirs():
        candidate = images_dir / Path(clean_relative)
        if candidate.exists():
            return candidate
    return None


def _versioned_image_path(image_name: str) -> str:
    relative_path = image_relative_path(image_name)
    if not relative_path:
        return ""

    image_path = _resolve_existing_image_path(relative_path)
    if not image_path:
        return ""

    version_suffix = f"?v={int(image_path.stat().st_mtime)}"
    return f"/images/{relative_path.replace('\\', '/')}{version_suffix}"


def _build_image_lookup() -> dict[str, list[tuple[str, str]]]:
    lookup: dict[str, list[tuple[str, str]]] = {"aquant": [], "kohler": []}
    seen: dict[str, set[str]] = {"aquant": set(), "kohler": set()}

    for images_dir in _candidate_images_dirs():
        for path in images_dir.rglob("*.png"):
            try:
                relative = path.relative_to(images_dir).as_posix()
            except ValueError:
                continue

            source_key = "kohler" if relative.startswith("Kohler/") else "aquant"
            if relative in seen[source_key]:
                continue
            seen[source_key].add(relative)
            lookup[source_key].append((normalize_code(path.stem), relative))

    for source_key in lookup:
        lookup[source_key].sort(key=lambda item: (len(item[0]), item[1]))
    return lookup


IMAGE_LOOKUP = _build_image_lookup()


def _fallback_image_by_code(code: str, source_key: str) -> str | None:
    compact_code = normalize_code(code)
    if not compact_code:
        return None

    source_images = IMAGE_LOOKUP.get(source_key, [])
    for stem_compact, relative in source_images:
        if stem_compact.startswith(compact_code):
            return relative
    return None


def _image_name_from_code(code: str) -> str:
    value = re.sub(r"\s*([+/\-])\s*", r"\1", str(code or "").strip().upper())
    value = value.replace("\\", "/")

    parts = []
    for raw_part in value.split("/"):
        part = raw_part.replace(" ", "").strip(".-")
        part = "".join("-" if char in WINDOWS_FORBIDDEN_SEGMENT_CHARS else char for char in part)
        if part:
            parts.append(part)

    safe = "/".join(parts).strip("/")
    return f"{safe}.png" if safe else ""


def _searchable_product(product: dict) -> dict:
    code_text = str(product.get("code", "")).strip().upper()
    return {
        **product,
        "_code_raw": code_text,
        "_code_tokens": [token for token in re.split(r"[+/]", code_text) if token],
        "_code_compact": normalize_code(product.get("code", "")),
        "_name_normalized": normalize_text(product.get("name", "")),
        "_name_compact": normalize_code(product.get("name", "")),
        "_details_normalized": normalize_text(product.get("details", "")),
        "_details_compact": normalize_code(product.get("details", "")),
    }


def _build_source_store(catalog: list[dict]) -> dict:
    searchable_catalog = [_searchable_product(product) for product in catalog]
    exact_code_index = {}

    for product in searchable_catalog:
        key = product["_code_compact"]
        existing = exact_code_index.get(key)
        if existing is None or len(product.get("name", "")) < len(existing.get("name", "")):
            exact_code_index[key] = product

    return {
        "catalog": catalog,
        "searchable": searchable_catalog,
        "exact": exact_code_index,
    }


def _is_catalog_suspicious(catalog: list[dict], source_key: str) -> bool:
    if not catalog:
        return True

    # Reject placeholder/minimal datasets that sometimes appear in deployment artifacts.
    if len(catalog) <= 2:
        source_token = normalize_text(source_key)
        placeholder_rows = 0
        for item in catalog:
            code_text = normalize_text(item.get("code", ""))
            name_text = normalize_text(item.get("name", ""))
            if code_text == source_token or name_text == source_token:
                placeholder_rows += 1
        if placeholder_rows >= 1:
            return True

    return False


def _split_code_variant(code: str) -> tuple[str, str]:
    text = str(code or "").strip().upper()
    number_match = re.search(r"(\d{3,5})", text)
    if not number_match:
        return "", ""

    base_code = number_match.group(1)
    tail = text[number_match.end() :].strip(" -:+/")
    tail = re.sub(r"[^A-Z0-9]+", "", tail)
    variant = tail[:6] if tail else ""
    return base_code, variant


def _coerce_price(value) -> int:
    """Normalize price values from excel/cache/json into a safe integer rupee amount."""
    if value is None:
        return 0

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        try:
            return max(0, int(round(float(value))))
        except (TypeError, ValueError):
            return 0

    text = str(value).strip()
    if not text:
        return 0

    # Handle common production payload formats such as "₹ 12,345/-".
    cleaned = text.replace(",", "")
    match = re.search(r"\d+(?:\.\d{1,2})?", cleaned)
    if not match:
        return 0

    try:
        return max(0, int(round(float(match.group(0)))))
    except (TypeError, ValueError):
        return 0


def _infer_variant_from_color(color: str) -> str:
    normalized_color = normalize_text(color).strip()
    if not normalized_color:
        return ""

    for variant, variant_color in AQUANT_VARIANT_COLOR_MAP.items():
        if normalize_text(variant_color) == normalized_color:
            return variant
    return ""


def _load_catalog_from_excel(excel_path: Path, source_key: str, source_label: str) -> list[dict]:
    if not excel_path.exists():
        return []

    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as error:
        print(f"[catalog:{source_key}] failed to read excel: {error}")
        return []

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []

        fields = [str(value).strip().lower() if value is not None else "" for value in header]
        index_map = {name: idx for idx, name in enumerate(fields) if name}

        def _value(row: tuple, key: str, default=""):
            idx = index_map.get(key)
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return default if value is None else value

        catalog: list[dict] = []
        seen_codes = set()

        for row in rows:
            code = str(_value(row, "code", "")).strip()
            name = str(_value(row, "name", "")).strip()
            if not code or not name:
                continue

            compact_code = normalize_code(code)
            if not compact_code or compact_code in seen_codes:
                continue

            seen_codes.add(compact_code)

            price = _coerce_price(_value(row, "price", 0))

            image_file = str(_value(row, "image_file", "")).strip()
            image_value = str(_value(row, "image", "")).strip()
            if image_file:
                image_value = _versioned_image_path(image_file)

            try:
                page_number = int(float(_value(row, "page_number", 0) or 0))
            except (TypeError, ValueError):
                page_number = 0

            source_value = str(_value(row, "source", source_key)).strip().lower() or source_key
            source_label_value = str(_value(row, "source_label", source_label)).strip() or source_label

            base_code_value = str(_value(row, "base_code", "")).strip()
            variant_value = str(_value(row, "variant", "")).strip().upper()
            if not base_code_value:
                parsed_base, parsed_variant = _split_code_variant(code)
                base_code_value = parsed_base
                if not variant_value:
                    variant_value = parsed_variant
            is_cp_value = str(_value(row, "is_cp", "")).strip()
            is_cp = is_cp_value in {"1", "true", "True", "yes", "YES"} or variant_value == "CP"

            details = str(_value(row, "details", name)).strip() or name
            color = str(_value(row, "color", "")).strip() or None
            size = str(_value(row, "size", "")).strip() or None
            if not variant_value and color:
                inferred_variant = _infer_variant_from_color(color)
                if inferred_variant:
                    variant_value = inferred_variant
                    is_cp = variant_value == "CP"
            if not color and variant_value in AQUANT_VARIANT_COLOR_MAP:
                color = AQUANT_VARIANT_COLOR_MAP[variant_value]
            is_cp = is_cp or variant_value == "CP"

            catalog.append(
                {
                    "source": source_value,
                    "source_label": source_label_value,
                    "code": code,
                    "name": name,
                    "price": price,
                    "color": color,
                    "size": size,
                    "details": details,
                    "page_number": page_number,
                    "image": image_value,
                    "image_bbox": None,
                    "base_code": base_code_value or None,
                    "variant": variant_value or None,
                    "is_cp": is_cp,
                }
            )

        catalog.sort(key=lambda item: (item["code"], item["name"]))
        return catalog
    finally:
        workbook.close()


def _load_catalog_from_cache(cache_path: Path, source_key: str, source_label: str) -> list[dict]:
    if not cache_path.exists():
        return []

    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        print(f"[catalog:{source_key}] failed to read cache: {error}")
        return []

    if not isinstance(data, list):
        return []

    catalog: list[dict] = []
    seen_codes = set()

    for item in data:
        if not isinstance(item, dict):
            continue

        code = str(item.get("code", "")).strip()
        name = str(item.get("name", "")).strip()
        if not code or not name:
            continue

        compact_code = normalize_code(code)
        if not compact_code or compact_code in seen_codes:
            continue

        seen_codes.add(compact_code)

        price = _coerce_price(item.get("price", 0))

        source_value = str(item.get("source", source_key)).strip().lower() or source_key
        source_label_value = str(item.get("source_label") or item.get("sourceLabel") or source_label).strip() or source_label

        base_code_value = str(item.get("base_code") or item.get("baseCode") or "").strip()
        variant_value = str(item.get("variant", "")).strip().upper()
        if not base_code_value:
            parsed_base, parsed_variant = _split_code_variant(code)
            base_code_value = parsed_base
            if not variant_value:
                variant_value = parsed_variant

        is_cp_value = str(item.get("is_cp") or item.get("isCp") or "").strip()
        is_cp = is_cp_value in {"1", "true", "True", "yes", "YES"} or variant_value == "CP"

        color = str(item.get("color", "")).strip() or None
        if not variant_value and color:
            inferred_variant = _infer_variant_from_color(color)
            if inferred_variant:
                variant_value = inferred_variant
                is_cp = variant_value == "CP"
        if not color and variant_value in AQUANT_VARIANT_COLOR_MAP:
            color = AQUANT_VARIANT_COLOR_MAP[variant_value]
        is_cp = is_cp or variant_value == "CP"

        image_value = str(item.get("image", "")).strip()
        if image_value:
            image_value = _versioned_image_path(image_value)

        try:
            page_number = int(float(item.get("page_number") or item.get("pageNumber") or 0))
        except (TypeError, ValueError):
            page_number = 0

        catalog.append(
            {
                "source": source_value,
                "source_label": source_label_value,
                "code": code,
                "name": name,
                "price": price,
                "color": color,
                "size": str(item.get("size", "")).strip() or None,
                "details": str(item.get("details", name)).strip() or name,
                "page_number": page_number,
                "image": image_value,
                "image_bbox": item.get("image_bbox") or item.get("imageBbox"),
                "base_code": base_code_value or None,
                "variant": variant_value or None,
                "is_cp": is_cp,
            }
        )

    catalog.sort(key=lambda product: (product["code"], product["name"]))
    return catalog


def _load_catalog_from_products_file(products_path: Path, source_key: str, source_label: str) -> list[dict]:
    if not products_path.exists():
        return []

    try:
        data = json.loads(products_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        print(f"[catalog:{source_key}] failed to read products fallback: {error}")
        return []

    if not isinstance(data, list):
        return []

    catalog: list[dict] = []
    seen_rows = set()

    for item in data:
        if not isinstance(item, dict):
            continue

        explicit_source = str(item.get("source", "")).strip().lower()
        if explicit_source and explicit_source != source_key:
            continue
        if not explicit_source and source_key != "aquant":
            # Legacy source-less fallback files are Aquant-only; don't mirror into Kohler.
            continue
        source_value = explicit_source or "aquant"

        code = str(item.get("code", "")).strip()
        name = str(item.get("name") or item.get("product_name") or "").strip()
        if not code:
            continue
        if not name:
            name = code

        price = _coerce_price(item.get("price", 0))

        base_code_value = str(item.get("base_code") or item.get("baseCode") or "").strip()
        variant_value = str(item.get("variant", "")).strip().upper()
        if not base_code_value:
            parsed_base, parsed_variant = _split_code_variant(code)
            base_code_value = parsed_base
            if not variant_value:
                variant_value = parsed_variant

        color = str(item.get("color") or item.get("finish") or "").strip() or None
        size_value = str(item.get("size", "")).strip() or None
        if not variant_value and color:
            inferred_variant = _infer_variant_from_color(color)
            if inferred_variant:
                variant_value = inferred_variant
        if not color and variant_value in AQUANT_VARIANT_COLOR_MAP:
            color = AQUANT_VARIANT_COLOR_MAP[variant_value]

        image_value = str(item.get("image") or item.get("image_file") or "").strip()
        if image_value:
            image_value = _versioned_image_path(image_value)

        try:
            page_number = int(float(item.get("page_number") or item.get("pageNumber") or 0))
        except (TypeError, ValueError):
            page_number = 0

        is_cp_value = str(item.get("is_cp") or item.get("isCp") or "").strip()
        is_cp = is_cp_value in {"1", "true", "True", "yes", "YES"} or variant_value == "CP"

        details = str(item.get("details") or name).strip() or name
        source_label_value = str(item.get("source_label") or item.get("sourceLabel") or source_label).strip() or source_label

        dedupe_key = (
            normalize_code(code),
            normalize_text(name),
            normalize_text(details),
            normalize_text(color or ""),
            normalize_text(size_value or ""),
            normalize_code(variant_value),
            str(price),
        )
        if dedupe_key in seen_rows:
            continue
        seen_rows.add(dedupe_key)

        catalog.append(
            {
                "source": source_value,
                "source_label": source_label_value,
                "code": code,
                "name": name,
                "price": price,
                "color": color,
                "size": size_value,
                "details": details,
                "page_number": page_number,
                "image": image_value,
                "image_bbox": item.get("image_bbox") or item.get("imageBbox"),
                "base_code": base_code_value or None,
                "variant": variant_value or None,
                "is_cp": is_cp,
            }
        )

    catalog.sort(key=lambda product: (product["code"], product["name"]))
    return catalog


def _resolve_excel_path(excel_path: Path) -> Path:
    candidates = sorted(
        excel_path.parent.glob(f"{excel_path.stem}_codeimg*{excel_path.suffix}"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]

    codeimg_path = excel_path.with_name(f"{excel_path.stem}_codeimg{excel_path.suffix}")
    if codeimg_path.exists():
        return codeimg_path
    return excel_path


def _log_kohler_runtime_paths() -> None:
    print("Kohler file exists:", os.path.exists(KOHLER_CACHE_PATH))
    print("Kohler excel exists:", os.path.exists(KOHLER_EXCEL_PATH))
    print("Current directory:", os.getcwd())
    try:
        print("Files:", os.listdir(BASE_DIR))
    except OSError as error:
        print("Files: <error>", error)


SUPABASE_IMAGE_CACHE: dict[str, str] = {}

def load_catalogs() -> dict[str, dict]:
    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_KEY")
    if supabase_url and supabase_key:
        print("[startup] loading from Supabase")
        import requests
        try:
            res = requests.get(
                f"{supabase_url}/rest/v1/products?limit=5000",
                headers={"apikey": supabase_key, "Authorization": f"Bearer {supabase_key}"}
            )
            if res.status_code == 200:
                data = res.json()
                aquant = [p for p in data if p.get('source') == 'aquant']
                kohler = [p for p in data if p.get('source') == 'kohler']
                
                try:
                    list_res = requests.post(
                        f"{supabase_url}/storage/v1/object/list/product-images",
                        headers={"Authorization": f"Bearer {supabase_key}"},
                        json={"limit": 10000}
                    )
                    if list_res.status_code == 200:
                        for obj in list_res.json():
                            name = obj.get("name")
                            if name:
                                SUPABASE_IMAGE_CACHE[normalize_code(name)] = name
                                
                    list_res_kohler = requests.post(
                        f"{supabase_url}/storage/v1/object/list/product-images",
                        headers={"Authorization": f"Bearer {supabase_key}"},
                        json={"prefix": "Kohler/", "limit": 10000}
                    )
                    if list_res_kohler.status_code == 200:
                        for obj in list_res_kohler.json():
                            name = obj.get("name")
                            if name:
                                SUPABASE_IMAGE_CACHE[normalize_code(name)] = f"Kohler/{name}"
                except Exception as e:
                    print("[startup] Supabase image list failed:", e)

                return {
                    "aquant": _build_source_store(aquant),
                    "kohler": _build_source_store(kohler)
                }
        except Exception as e:
            print("[startup] Supabase load failed:", e)

    source_store = {}
    allow_kohler_products_fallback = os.environ.get("ENABLE_KOHLER_PRODUCTS_FALLBACK", "").strip().lower() in {
        "1",
        "true",
        "yes",
    }

    _log_kohler_runtime_paths()
    print("use_cache:", USE_CACHE)

    for source_key, config in CATALOG_SOURCES.items():
        excel_path = Path(config.get("excel_path", "")) if config.get("excel_path") else None
        cache_path = Path(config.get("cache_path", "")) if config.get("cache_path") else None

        candidates: list[tuple[str, list[dict], str]] = []

        if excel_path:
            excel_path = _resolve_excel_path(excel_path)
            excel_catalog = _load_catalog_from_excel(
                excel_path=excel_path,
                source_key=source_key,
                source_label=config["label"],
            )
            if excel_catalog and not _is_catalog_suspicious(excel_catalog, source_key):
                candidates.append(("excel", excel_catalog, excel_path.name))
            elif excel_catalog:
                print(
                    f"[catalog:{source_key}] ignoring suspicious excel dataset ({len(excel_catalog)}) from {excel_path.name}"
                )

        # Products fallback can contain stale Kohler rows on deployment;
        # keep it disabled for Kohler unless explicitly enabled.
        if source_key != "kohler" or allow_kohler_products_fallback:
            for products_path in FALLBACK_PRODUCTS_PATHS:
                fallback_catalog = _load_catalog_from_products_file(
                    products_path=products_path,
                    source_key=source_key,
                    source_label=config["label"],
                )
                if fallback_catalog:
                    candidates.append(("fallback", fallback_catalog, products_path.name))

        if USE_CACHE and cache_path:
            cache_catalog = _load_catalog_from_cache(
                cache_path=cache_path,
                source_key=source_key,
                source_label=config["label"],
            )
            if cache_catalog:
                candidates.append(("cache", cache_catalog, cache_path.name))

        if not candidates:
            print(f"[catalog:{source_key}] no valid data source found")
            source_store[source_key] = _build_source_store([])
            continue

        # Kohler is safer from excel first because stale deployment caches can carry zero-price rows.
        if source_key == "kohler":
            source_priority = {"excel": 0, "cache": 1, "fallback": 2}
        else:
            source_priority = {"cache": 0, "excel": 1, "fallback": 2}
        candidates.sort(key=lambda item: (-len(item[1]), source_priority.get(item[0], 9), item[2]))
        selected_kind, selected_catalog, selected_origin = candidates[0]

        # Backfill zero or missing prices from other candidate sources using exact code match.
        replacement_price_by_code: dict[str, int] = {}
        for _, candidate_catalog, _ in candidates:
            for row in candidate_catalog:
                code_key = normalize_code(row.get("code", ""))
                if not code_key:
                    continue
                candidate_price = _coerce_price(row.get("price", 0))
                if candidate_price > 0 and candidate_price > replacement_price_by_code.get(code_key, 0):
                    replacement_price_by_code[code_key] = candidate_price

        repaired_count = 0
        for row in selected_catalog:
            code_key = normalize_code(row.get("code", ""))
            if not code_key:
                continue
            current_price = _coerce_price(row.get("price", 0))
            replacement_price = replacement_price_by_code.get(code_key, 0)
            if current_price <= 0 and replacement_price > 0:
                row["price"] = replacement_price
                repaired_count += 1

        print(
            f"[catalog:{source_key}] selected {selected_kind} source {selected_origin} with {len(selected_catalog)} products"
            + (f" (price backfilled: {repaired_count})" if repaired_count else "")
        )
        source_store[source_key] = _build_source_store(selected_catalog)

    return source_store


def _catalog_sources_signature() -> tuple:
    """Build a lightweight signature of catalog source files to detect updates."""
    entries: list[tuple[str, str, int, int]] = []

    for source_key, config in CATALOG_SOURCES.items():
        excel_path = Path(config.get("excel_path", "")) if config.get("excel_path") else None
        cache_path = Path(config.get("cache_path", "")) if USE_CACHE and config.get("cache_path") else None

        if excel_path:
            resolved_excel = _resolve_excel_path(excel_path)
            try:
                stats = resolved_excel.stat()
                entries.append((source_key, f"excel:{resolved_excel.name}", stats.st_mtime_ns, stats.st_size))
            except OSError:
                entries.append((source_key, f"excel:{resolved_excel.name}", -1, -1))

        if cache_path:
            try:
                stats = cache_path.stat()
                entries.append((source_key, f"cache:{cache_path.name}", stats.st_mtime_ns, stats.st_size))
            except OSError:
                entries.append((source_key, f"cache:{cache_path.name}", -1, -1))

    for fallback_path in FALLBACK_PRODUCTS_PATHS:
        try:
            stats = fallback_path.stat()
            entries.append(("fallback", str(fallback_path), stats.st_mtime_ns, stats.st_size))
        except OSError:
            entries.append(("fallback", str(fallback_path), -1, -1))

    return tuple(entries)


SOURCE_STORE = load_catalogs()
_CATALOG_SOURCES_SIGNATURE = _catalog_sources_signature()


@app.on_event("startup")
def _startup_load_catalogs() -> None:
    global SOURCE_STORE, _CATALOG_SOURCES_SIGNATURE

    print("[startup] loading catalogs")
    _log_kohler_runtime_paths()
    SOURCE_STORE = load_catalogs()
    _CATALOG_SOURCES_SIGNATURE = _catalog_sources_signature()
    print("[startup] catalog counts:", {source_key: len(store.get("catalog", [])) for source_key, store in SOURCE_STORE.items()})


def _ensure_catalogs_loaded() -> None:
    """Reload catalog data when any source file has changed on disk."""
    global SOURCE_STORE, _CATALOG_SOURCES_SIGNATURE

    latest_signature = _catalog_sources_signature()
    if latest_signature == _CATALOG_SOURCES_SIGNATURE:
        return

    SOURCE_STORE = load_catalogs()
    _CATALOG_SOURCES_SIGNATURE = latest_signature


def _fallback_files_status() -> dict[str, bool]:
    status: dict[str, bool] = {}
    for path in FALLBACK_PRODUCTS_PATHS:
        try:
            resolved = path.resolve()
        except OSError:
            resolved = path
        status[str(resolved)] = resolved.exists()
    return status


def _is_probably_code_query(query: str) -> bool:
    trimmed = query.strip()
    if not trimmed or not any(character.isdigit() for character in trimmed):
        return False
    return bool(re.fullmatch(r"[0-9A-Za-z+\- /]+", trimmed))


def _query_code_tokens(query: str) -> list[str]:
    cleaned = str(query or "").strip().upper()
    if not cleaned:
        return []
    return [token for token in re.split(r"[+]", cleaned) if token.strip()]


def _relaxed_code_queries(query: str) -> list[str]:
    text = str(query or "").strip().upper()
    if not text:
        return []

    compact = normalize_code(text).upper()
    match = re.match(r"^(\d{4,7})([A-Z].*)?$", compact)
    if not match:
        return []

    digits = match.group(1)
    suffix = match.group(2) or ""
    relaxed = []

    for index, character in enumerate(digits):
        if character != "0":
            continue
        candidate_digits = digits[:index] + digits[index + 1 :]
        if 3 <= len(candidate_digits) <= 5:
            candidate = f"{candidate_digits}{suffix}"
            if candidate != compact and candidate not in relaxed:
                relaxed.append(candidate)

    return relaxed


def _kohler_code_alias_queries(query: str) -> list[str]:
    """Generate Kohler code aliases for common typing/finish variations."""
    text = str(query or "").strip().upper().replace(" ", "")
    if not text:
        return []

    aliases: list[str] = []

    def _add(candidate: str) -> None:
        candidate = candidate.strip()
        if candidate and candidate != text and candidate not in aliases:
            aliases.append(candidate)

    # Users often type an extra EX prefix; try canonical Kohler K- form too.
    if text.startswith("EX") and len(text) > 2:
        _add(text[2:])
        _add(f"K-{text[2:]}")

    # Common finish mismatch: BRD is often catalogued as RGD or as base code.
    if text.endswith("-BRD"):
        _add(text[:-4])
        _add(f"{text[:-4]}-RGD")

    return aliases


def _combined_code_search(query: str, source_key: str) -> list[dict]:
    if source_key not in SOURCE_STORE:
        return []

    source = SOURCE_STORE[source_key]
    compact_query = normalize_code(query)
    exact_combined = source["exact"].get(compact_query)
    if exact_combined is not None:
        return [exact_combined]

    for relaxed_query in _relaxed_code_queries(query):
        exact_relaxed = source["exact"].get(normalize_code(relaxed_query))
        if exact_relaxed is not None:
            return [exact_relaxed]

    searchable = source["searchable"]
    tokens = _query_code_tokens(query)
    if len(tokens) < 2:
        return []

    # Ignore ambiguous non-numeric tokens (for example: W, WN) to avoid
    # accidental cross-product combinations like 1871 + 1021W.
    resolvable_tokens = [token for token in tokens if any(character.isdigit() for character in token)]
    if not resolvable_tokens:
        return []

    matched_products: list[dict] = []
    for token in resolvable_tokens:
        compact = normalize_code(token)
        if not compact:
            continue

        exact = source["exact"].get(compact)
        if exact is not None:
            matched_products.append(exact)
            continue

        fallback = next(
            (
                product
                for product in searchable
                if product["_code_compact"].startswith(compact)
                or any(normalize_code(piece) == compact for piece in product.get("_code_tokens", []))
            ),
            None,
        )
        if fallback is not None:
            matched_products.append(fallback)

    unique_products = []
    seen = set()
    for product in matched_products:
        key = normalize_code(product.get("code", ""))
        if key in seen:
            continue
        seen.add(key)
        unique_products.append(product)

    if len(unique_products) < 2:
        # For mixed queries such as "1871AB+WN", fall back to the first
        # resolvable product instead of manufacturing a wrong combination.
        return unique_products[:1]

    combined_name = " + ".join(str(product.get("name", "")).strip() for product in unique_products if product.get("name"))
    primary = unique_products[0]
    combined_result = {
        **primary,
        "code": " + ".join(str(product.get("code", "")).strip() for product in unique_products),
        "name": combined_name or primary.get("name"),
        "details": " + ".join(str(product.get("details", "")).strip() for product in unique_products if product.get("details")),
        "price": int(primary.get("price", 0) or 0),
        "is_cp": any(bool(product.get("is_cp")) for product in unique_products),
        "combined_products": unique_products,
    }
    return [combined_result]


def _levenshtein_distance(s1: str, s2: str) -> int:
    """Calculate Levenshtein distance for fuzzy matching."""
    if len(s1) < len(s2):
        return _levenshtein_distance(s2, s1)
    
    if len(s2) == 0:
        return len(s1)
    
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
    
    return previous_row[-1]


def _fuzzy_match_score(query: str, target: str) -> float:
    """Score fuzzy match between 0-100. Higher is better."""
    query = query.strip().lower()
    target = target.strip().lower()
    
    if not query or not target:
        return 0.0
    
    # Exact match
    if query == target:
        return 100.0
    
    # Starts with
    if target.startswith(query):
        return 90.0 + (10.0 * (1.0 - len(query) / len(target)))
    
    # Contains as substring
    if query in target:
        return 70.0
    
    # Fuzzy: Levenshtein distance
    max_len = max(len(query), len(target))
    distance = _levenshtein_distance(query, target)
    similarity = 1.0 - (distance / max_len)
    
    if similarity >= 0.6:
        return similarity * 80.0
    
    return 0.0


def _get_autocomplete_suggestions(query: str, source_key: str, limit: int = 10) -> list[dict]:
    """Get autocomplete suggestions based on prefix matching."""
    query = query.strip()
    if not query or len(query) < 1:
        return []
    
    if source_key not in SOURCE_STORE:
        return []
    
    source = SOURCE_STORE[source_key]
    searchable = source["searchable"]
    compact_query = normalize_code(query)
    is_code_query = _is_probably_code_query(query)
    
    suggestions = []
    seen_codes = set()
    
    for product in searchable:
        if product["code"] in seen_codes:
            continue
        
        code_compact = product["_code_compact"]
        name_normalized = product["_name_normalized"]
        
        score = 0.0
        
        if is_code_query:
            # Code-based autocomplete
            if code_compact.startswith(compact_query):
                score = 95.0
            elif compact_query in code_compact:
                score = 80.0
        else:
            # Name-based autocomplete
            if name_normalized.startswith(query.lower()):
                score = 90.0
            elif query.lower() in name_normalized:
                score = 70.0
        
        if score > 0:
            suggestions.append((score, product))
            seen_codes.add(product["code"])
    
    # Sort by score (descending) and return top results
    suggestions.sort(key=lambda x: -x[0])
    if suggestions:
        return [item[1] for item in suggestions[:limit]]

    image_only_results = _image_only_query_results(query, source_key)
    if image_only_results:
        image_only_product = image_only_results[0]
        return [
            {
                "code": image_only_product.get("code", ""),
                "name": image_only_product.get("name", ""),
                "source": image_only_product.get("source", ""),
            }
        ]

    return []


def _search_matches(query: str, source_key: str, limit: int = 20) -> list[dict]:
    query = query.strip()
    if not query or source_key not in SOURCE_STORE:
        return []

    normalized_query = normalize_text(query)
    compact_query = normalize_code(query)
    source = SOURCE_STORE[source_key]
    searchable = source["searchable"]

    combined_hits = _combined_code_search(query, source_key)
    if combined_hits:
        return combined_hits

    if _is_probably_code_query(query):
        if len(compact_query) < 2:
            return []

        code_candidates = []
        for product in searchable:
            code_compact = product["_code_compact"]
            score = 0
            if code_compact == compact_query:
                score = 320
            elif normalize_code(product.get("_code_raw", "")) == compact_query:
                score = 300
            elif code_compact.startswith(compact_query):
                score = 260
            elif compact_query in code_compact:
                score = 200
            elif any(normalize_code(token) == compact_query for token in product.get("_code_tokens", [])):
                score = 215

            if score:
                code_candidates.append(
                    (
                        score,
                        0 if product.get("image") else 1,
                        0 if product.get("color") else 1,
                        len(product.get("code", "")),
                        len(product.get("name", "")),
                        product,
                    )
                )

        if not code_candidates:
            for relaxed_query in _relaxed_code_queries(query):
                exact_relaxed = source["exact"].get(normalize_code(relaxed_query))
                if exact_relaxed is not None:
                    return [exact_relaxed]

            if source_key == "kohler":
                for alias_query in _kohler_code_alias_queries(query):
                    exact_alias = source["exact"].get(normalize_code(alias_query))
                    if exact_alias is not None:
                        return [exact_alias]

                    alias_compact = normalize_code(alias_query)
                    if not alias_compact:
                        continue
                    alias_candidates = [
                        (
                            score,
                            0 if product.get("image") else 1,
                            0 if product.get("color") else 1,
                            len(product.get("code", "")),
                            len(product.get("name", "")),
                            product,
                        )
                        for product in searchable
                        for score in [
                            320 if product["_code_compact"] == alias_compact
                            else 300 if normalize_code(product.get("_code_raw", "")) == alias_compact
                            else 260 if product["_code_compact"].startswith(alias_compact)
                            else 200 if alias_compact in product["_code_compact"]
                            else 0
                        ]
                        if score
                    ]
                    if alias_candidates:
                        alias_candidates.sort(key=lambda item: (-item[0], item[1], item[2], item[3], item[4]))
                        return [item[-1] for item in alias_candidates[:limit]]

            if compact_query.isdigit() and len(compact_query) >= 5:
                relaxed_queries = []
                if compact_query.startswith("1"):
                    relaxed_queries.append(compact_query[1:])
                relaxed_queries.append(compact_query[-4:])

                seen_relaxed = set()
                for relaxed_query in relaxed_queries:
                    relaxed_query = relaxed_query.strip()
                    if len(relaxed_query) < 3 or relaxed_query in seen_relaxed:
                        continue
                    seen_relaxed.add(relaxed_query)

                    relaxed_candidates = []
                    for product in searchable:
                        code_compact = product["_code_compact"]
                        score = 0
                        if code_compact == relaxed_query:
                            score = 320
                        elif normalize_code(product.get("_code_raw", "")) == relaxed_query:
                            score = 300
                        elif code_compact.startswith(relaxed_query):
                            score = 260
                        elif relaxed_query in code_compact:
                            score = 200
                        elif any(normalize_code(token) == relaxed_query for token in product.get("_code_tokens", [])):
                            score = 215

                        if score:
                            relaxed_candidates.append(
                                (
                                    score,
                                    0 if product.get("image") else 1,
                                    0 if product.get("color") else 1,
                                    len(product.get("code", "")),
                                    len(product.get("name", "")),
                                    product,
                                )
                            )

                    if relaxed_candidates:
                        relaxed_candidates.sort(key=lambda item: (-item[0], item[1], item[2], item[3], item[4]))
                        return [item[-1] for item in relaxed_candidates[:limit]]

            return []

        code_candidates.sort(key=lambda item: (-item[0], item[1], item[2], item[3], item[4]))
        return [item[-1] for item in code_candidates[:limit]]

    if compact_query in source["exact"]:
        return [source["exact"][compact_query]]

    for relaxed_query in _relaxed_code_queries(query):
        exact_relaxed = source["exact"].get(normalize_code(relaxed_query))
        if exact_relaxed is not None:
            return [exact_relaxed]

    candidates = []
    for product in searchable:
        score = 0
        code_compact = product["_code_compact"]
        name_normalized = product["_name_normalized"]
        name_compact = product["_name_compact"]
        details_normalized = product["_details_normalized"]
        details_compact = product["_details_compact"]

        if compact_query and code_compact.startswith(compact_query):
            score = max(score, 120)
        if compact_query and code_compact == compact_query:
            score = max(score, 140)
        if compact_query and compact_query in code_compact:
            score = max(score, 110)
        if normalized_query and name_normalized == normalized_query:
            score = max(score, 100)
        if normalized_query and name_normalized.startswith(normalized_query):
            score = max(score, 90)
        if normalized_query and normalized_query in name_normalized:
            score = max(score, 80)
        if normalized_query and normalized_query in details_normalized:
            score = max(score, 75)
        if compact_query and compact_query in details_compact:
            score = max(score, 78)
        if compact_query and compact_query in name_compact:
            score = max(score, 70)

        if score:
            candidates.append(
                (
                    score,
                    0 if product.get("image") else 1,
                    0 if product.get("color") else 1,
                    abs(len(name_normalized) - len(normalized_query)),
                    len(product.get("name", "")),
                    product,
                )
            )

    if not candidates:
        return []

    candidates.sort(key=lambda item: (-item[0], item[1], item[2], item[3], item[4]))
    return [item[-1] for item in candidates[:limit]]


def _manual_query_results(query: str, source_key: str) -> list[dict]:
    if source_key != "aquant":
        return []

    query_key = normalize_code(query)
    items = MANUAL_QUERY_RESULTS.get(query_key, [])
    if not items:
        return []

    results = []
    for item in items:
        results.append(
            {
                "source": "aquant",
                "source_label": "Aquant",
                "code": item["code"],
                "name": item.get("name") or "Product",
                "price": item["price"],
                "color": item["color"],
                "size": item.get("size"),
                "details": item.get("details") or item.get("name"),
                "base_code": item.get("base_code") or normalize_code(item["code"]),
                "variant": item.get("variant"),
                "is_cp": False,
                "image": item.get("image"),
            }
        )
    return results


def _image_only_query_results(query: str, source_key: str) -> list[dict]:
    code_value = str(query or "").strip()
    if not code_value:
        return []

    compact_code = normalize_code(code_value)
    if not compact_code:
        return []

    source = SOURCE_STORE.get(source_key)
    if source and source.get("exact", {}).get(compact_code):
        # Prefer real catalog rows (with actual price/details) when they exist.
        return []

    expected_image_file = _image_name_from_code(code_value)
    candidates = []

    if expected_image_file:
        candidates.append(f"Kohler/{expected_image_file}" if source_key == "kohler" else expected_image_file)

    fallback_relative = next(
        (
            relative
            for stem_compact, relative in IMAGE_LOOKUP.get(source_key, [])
            if stem_compact == compact_code
        ),
        None,
    )
    if fallback_relative:
        candidates.append(fallback_relative)

    for relative_path in candidates:
        if not relative_path or not _resolve_existing_image_path(relative_path):
            continue

        base_code, variant = _split_code_variant(code_value)
        return [
            {
                "source": source_key,
                "source_label": CATALOG_SOURCES[source_key]["label"],
                "code": code_value,
                "name": code_value,
                "price": 0,
                "color": None,
                "size": None,
                "details": f"Manually added image for {code_value}",
                "base_code": base_code or normalize_code(code_value),
                "variant": variant or None,
                "is_cp": False,
                "image": f"/images/{relative_path.replace('\\', '/')}",
            }
        ]

    return []


def _serialize_product(request: Request, product: dict) -> dict:
    source_key = product.get("source") or "aquant"
    code_value = str(product.get("code") or "").strip()
    compact_code_value = normalize_code(code_value)

    supabase_url = os.environ.get("SUPABASE_URL")
    
    if supabase_url:
        if source_key == "kohler" and compact_code_value in KOHLER_NO_IMAGE_CODES:
            image = None
        else:
            db_image = str(product.get("image") or "").strip()
            if db_image:
                relative = image_relative_path(db_image)
                exact = SUPABASE_IMAGE_CACHE.get(normalize_code(relative))
                import urllib.parse
                if exact:
                    image = f"{supabase_url}/storage/v1/object/public/product-images/{urllib.parse.quote(exact)}"
                else:
                    # Fallback to the relative path, removing spaces (which is a common mismatch issue)
                    fallback_relative = relative.replace(" ", "")
                    image = f"{supabase_url}/storage/v1/object/public/product-images/{urllib.parse.quote(fallback_relative)}"
            else:
                image = None
    else:
        if source_key == "kohler" and compact_code_value in KOHLER_NO_IMAGE_CODES:
            image = None
        else:
            image = None
    
            raw_product_image = str(product.get("image") or "").strip()
            if raw_product_image:
                relative = image_relative_path(raw_product_image)
                if source_key == "kohler" and relative and not relative.startswith("Kohler/"):
                    relative = f"Kohler/{relative}"
                if relative and _resolve_existing_image_path(relative):
                    image = f"/images/{relative}"
    
            if not image:
                expected_image_file = _image_name_from_code(code_value)
                if expected_image_file:
                    if source_key == "kohler":
                        kohler_relative = f"Kohler/{expected_image_file}"
                        expected_image_path = _resolve_existing_image_path(kohler_relative)
                        if expected_image_path:
                            image = f"/images/{kohler_relative}"
                    else:
                        expected_image_path = _resolve_existing_image_path(expected_image_file)
                        if expected_image_path:
                            image = f"/images/{expected_image_file}"
    
            if not image:
                fallback_relative = _fallback_image_by_code(code_value, source_key)
                if fallback_relative and _resolve_existing_image_path(fallback_relative):
                    image = f"/images/{fallback_relative}"
    
            if image and str(image).startswith("/"):
                relative_image = str(image).split("?", 1)[0]
                if relative_image.startswith("/images/"):
                    relative_image = relative_image.removeprefix("/images/")
    
                image_path = _resolve_existing_image_path(relative_image)
                if image_path:
                    version = f"?v={int(image_path.stat().st_mtime)}"
                    image = f"{str(request.base_url).rstrip('/')}/images/{relative_image}{version}"
    
            if source_key == "kohler" and image:
                kohler_relative = image_relative_path(image)
                if not kohler_relative.startswith("Kohler/") or not _resolve_existing_image_path(kohler_relative):
                    image = None

    base_code_value = product.get("base_code")
    has_image = bool(image)

    serialized = {
        "source": source_key,
        "sourceLabel": product.get("source_label") or CATALOG_SOURCES[source_key]["label"],
        "name": product.get("name", ""),
        "code": product.get("code", ""),
        "baseCode": base_code_value,
        "groupId": str(base_code_value or normalize_code(product.get("code", "")) or "ungrouped"),
        "variant": product.get("variant"),
        "isCp": bool(product.get("is_cp")),
        "price": _coerce_price(product.get("price", 0)),
        "color": product.get("color"),
        "size": product.get("size"),
        "details": product.get("details"),
        "image": image,
        "hasImage": has_image,
    }

    if source_key == "aquant":
        override = PRODUCT_OVERRIDES.get(normalize_code(serialized.get("code", "")))
        if override:
            for field in ("name", "price", "color", "size", "details"):
                if field in override:
                    serialized[field] = override[field]

            override_image = str(override.get("image") or "").strip()
            if override_image:
                relative = image_relative_path(override_image)
                if supabase_url:
                    exact = SUPABASE_IMAGE_CACHE.get(normalize_code(relative))
                    import urllib.parse
                    if exact:
                        serialized["image"] = f"{supabase_url}/storage/v1/object/public/product-images/{urllib.parse.quote(exact)}"
                    else:
                        fallback_relative = relative.replace(" ", "")
                        serialized["image"] = f"{supabase_url}/storage/v1/object/public/product-images/{urllib.parse.quote(fallback_relative)}"
                    serialized["hasImage"] = True
                else:
                    resolved = _resolve_existing_image_path(relative)
                    if resolved:
                        version = f"?v={int(resolved.stat().st_mtime)}"
                        serialized["image"] = f"{str(request.base_url).rstrip('/')}/images/{relative}{version}"
                        serialized["hasImage"] = True

    combined_products = product.get("combined_products")
    if isinstance(combined_products, list) and combined_products:
        serialized["linkedProducts"] = [
            {
                "code": item.get("code", ""),
                "name": item.get("name", ""),
                "variant": item.get("variant"),
                "price": _coerce_price(item.get("price", 0)),
                "color": item.get("color"),
            }
            for item in combined_products
        ]

    return serialized


@app.get("/images/{image_path:path}")
def serve_image(image_path: str):
    resolved = _resolve_existing_image_path(image_path)
    if not resolved:
        return Response(status_code=404)
    return FileResponse(resolved)


@app.get("/")
def root():
    return "API is working"


@app.get("/health")
@app.get("/api/health")
def health():
    _ensure_catalogs_loaded()

    price_diagnostics = {}
    for source_key, store in SOURCE_STORE.items():
        catalog = store.get("catalog", [])
        total = len(catalog)
        with_price = sum(1 for row in catalog if _coerce_price(row.get("price", 0)) > 0)
        missing = max(0, total - with_price)
        price_diagnostics[source_key] = {
            "total": total,
            "with_price": with_price,
            "missing_price": missing,
            "with_price_percent": round((with_price / total) * 100, 2) if total else 0.0,
        }

    return {
        "status": "ok",
        "catalogs": {
            source_key: len(store["catalog"])
            for source_key, store in SOURCE_STORE.items()
        },
        "price_diagnostics": price_diagnostics,
        "images_dir": str(IMAGES_DIR),
        "cwd": str(Path.cwd()),
        "base_dir": str(BASE_DIR),
        "fallback_files": _fallback_files_status(),
        "use_cache": USE_CACHE,
    }


@app.get("/search")
@app.get("/api/search")
def search(
    request: Request,
    q: str = Query(default=""),
    query: str = Query(default=""),
    catalog: str = Query(default="all"),
):
    _ensure_catalogs_loaded()
    effective_query = (q or query or "").strip()

    selected_catalog = catalog.strip().lower()
    if selected_catalog not in {"all", *CATALOG_SOURCES.keys()}:
        selected_catalog = "all"

    if selected_catalog == "all":
        source_keys = list(CATALOG_SOURCES.keys())
    else:
        source_keys = [selected_catalog]

    matches = []
    seen_keys = set()
    for source_key in source_keys:
        source_matches = (
            _manual_query_results(effective_query, source_key)
            or _search_matches(effective_query, source_key)
            or _image_only_query_results(effective_query, source_key)
        )
        for match in source_matches:
            unique_key = (
                source_key,
                normalize_code(match.get("code", "")),
                normalize_text(match.get("name", "")),
            )
            if unique_key in seen_keys:
                continue
            seen_keys.add(unique_key)
            matches.append(_serialize_product(request, match))

    compact_query = normalize_code(effective_query)
    if compact_query and ("woodenseatcover" in compact_query or "walnutcolour" in compact_query):
        synthetic_key = ("aquant", "woodenseatcover", "walnutcolour")
        if synthetic_key not in seen_keys:
            matches.append(
                {
                    "source": "aquant",
                    "sourceLabel": "Aquant",
                    "name": "Wooden Seat Cover",
                    "code": "WSC",
                    "price": 17500,
                    "color": "Walnut Colour",
                    "size": "-",
                    "details": "WalnutColour MRP: Rs. 17,500/-",
                    "image": f"{str(request.base_url).rstrip('/')}/images/1857-wn.png",
                }
            )
            seen_keys.add(synthetic_key)

    return {
        "results": matches[:50],
    }


@app.get("/autocomplete")
@app.get("/api/autocomplete")
def autocomplete(
    q: str = Query(default=""),
    query: str = Query(default=""),
    catalog: str = Query(default="all"),
    limit: int = Query(default=10, ge=1, le=20),
):
    """Get autocomplete suggestions based on prefix matching."""
    _ensure_catalogs_loaded()
    effective_query = (q or query or "").strip()

    selected_catalog = catalog.strip().lower()
    if selected_catalog not in {"all", *CATALOG_SOURCES.keys()}:
        selected_catalog = "all"
    
    if selected_catalog == "all":
        source_keys = list(CATALOG_SOURCES.keys())
    else:
        source_keys = [selected_catalog]
    
    suggestions = []
    seen_codes = set()
    
    for source_key in source_keys:
        source_suggestions = _get_autocomplete_suggestions(effective_query, source_key, limit)
        for suggestion in source_suggestions:
            code_key = normalize_code(suggestion.get("code", ""))
            if code_key not in seen_codes:
                suggestions.append({
                    "code": suggestion.get("code", ""),
                    "name": suggestion.get("name", ""),
                    "source": suggestion.get("source", ""),
                })
                seen_codes.add(code_key)
                if len(suggestions) >= limit:
                    break
        
        if len(suggestions) >= limit:
            break
    
    return {
        "suggestions": suggestions[:limit]
    }


@app.post("/generate-pdf")
async def generate_pdf(data: dict):
    try:
        from pdf_service import generate_professional_pdf
        buffer = generate_professional_pdf(data)
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=quotation.pdf"}
        )
    except Exception as e:
        print(f"PDF GENERATION ERROR: {e}")
        return Response(content=str(e), status_code=500)
